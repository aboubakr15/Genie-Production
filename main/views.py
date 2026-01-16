from io import BytesIO
from django.urls import reverse
from openpyxl.utils import get_column_letter
from django.shortcuts import render, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from leads.forms import UploadSheetsForm
from .utils import has_valid_contact, is_valid_phone_number, clean_company_name, filter_companies, get_string_value, get_lead_related_data
from .models import (Lead, Sheet, LeadContactNames, LeadEmails, LeadPhoneNumbers, Log, LeadsAverage, UserLeader, FilterWords,
                    FilterType, LeadsColors, SalesLog, LeadTerminationCode, Notification, TaskLog, TerminationCode)
from .forms import LeadForm, AutoFillForm, FilterWordsForm #UploadFilesForm # ImportSheetsForm  # For mass importing 
import os, logging, pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from datetime import datetime, timezone as dt_timezone, timedelta
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl.styles import PatternFill
from django.contrib.auth.decorators import user_passes_test, login_required
from main.custom_decorators import is_in_group
from django.http import JsonResponse
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Q
from django.contrib.auth.models import User, Group
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json, pytz
import random, string
from concurrent.futures import ThreadPoolExecutor
from django.db import transaction
from .tasks import process_sheet_task
from django.conf import settings
import os

def generate_random_string(length=5):
    """Generate a random string of fixed length."""
    letters = string.ascii_letters  # Use both uppercase and lowercase letters
    return ''.join(random.choice(letters) for i in range(length))

logger = logging.getLogger('custom')

def index(request):
    if request.user.is_authenticated:
        first_group = request.user.groups.first()
        if first_group:
            return redirect(f"/{first_group.name}")
    return redirect("main:login")

def login_view(request):
    context = {"error": ""}
    
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        try:
            if not username or not password:
                raise ValidationError("Username and password are required.")
            
            if len(username) < 3 or len(username) > 20:
                raise ValidationError("Username must be between 3 and 20 characters long.")
            
            if len(password) < 8:
                raise ValidationError("Password must be at least 8 characters long.")

        except ValidationError as e:
            context['error'] = str(e)
            return render(request, "main/login.html", context=context)

        user = authenticate(request, username=username, password=password)

        if user is not None:
            send_cb_date_notifications()

            group = user.groups.first()
            if group:
                login(request, user)
                Log.objects.create(message=f"User [{user.username}] Logged in successfully")
                return redirect(f"/{group.name}")
            else:
                context['error'] = 'Oops, User has no group. Please ask the admin to add you to one.'
        else:
            logger.warning(f"Failed login attempt for {username}")
            context['error'] = 'Wrong username or password.'

    return render(request, "main/login.html", context=context)


def logout_view(request):
    username = request.user.username  # Get username before logging out
    logout(request)
    Log.objects.create(message=f"User [{username}] has logged out of the system")
    return redirect("/login/")


def sheet_list(request):
    term = request.GET.get('q', '')
    sheets = Sheet.objects.filter(name__icontains=term)
    results = [{'id': sheet.id, 'text': sheet.name} for sheet in sheets]
    return JsonResponse({'items': results})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager") or is_in_group(user, "Sales_manager"))
def lead_details(request, pk):
    # Get the lead instance with prefetched related data to avoid N+1 queries
    lead = get_object_or_404(
        Lead.objects.prefetch_related('leadphonenumbers_set', 'leademails_set', 'leadcontactnames_set', 'sheets'),
        pk=pk
    )
    
    # Retrieve all related phone numbers, emails, and contact names (already prefetched)
    phone_numbers = lead.leadphonenumbers_set.all()
    emails = lead.leademails_set.all()
    contact_names = lead.leadcontactnames_set.all()

    # Use a set to remove duplicates based on value
    unique_phone_numbers = {pn.value: pn for pn in phone_numbers}.values()
    unique_emails = {em.value: em for em in emails}.values()
    unique_contact_names = {cn.value: cn for cn in contact_names}.values()
    
    # Retrieve sheets associated with the lead (already prefetched)
    sheets = lead.sheets.all()
    
    # Get the user's group name (optimized with select_related)
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.select_related().first().name

    # Determine the template path based on group
    if group_name == 'operations_manager':
        template_path = 'operations_manager/lead_details.html'
    elif group_name == 'operations_team_leader':
        template_path = 'operations_team_leader/lead_details.html'
    elif group_name == 'sales_manager':
        template_path = 'sales_manager/lead_details.html'
    else:
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Pass all the details to the template
    data = {
        'lead': lead,
        'phone_numbers': unique_phone_numbers,
        'emails': unique_emails,
        'contact_names': unique_contact_names,
        'sheets': sheets,
    }
    
    return render(request, template_path, data)


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def add_lead(request):
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name
        
    if request.method == 'POST':
        form = LeadForm(request.POST)
        if form.is_valid():
            lead = form.save(commit=False)
            sheets = form.cleaned_data.get('sheets')
            
            if not sheets:
                messages.error(request, "Please select at least one sheet.")
                return render(request, "main/manage_lead_update.html", {'form': form})
            
            lead.save()
            
            # Process phone numbers with time zones
            phone_data = form.cleaned_data.get('phone_numbers')
            if phone_data:
                for line in phone_data.split('\n'):
                    line = line.strip()
                    if line:
                        # Split by comma, but be careful of time zones with commas
                        parts = line.split(',', 1)  # Split into max 2 parts
                        if len(parts) == 2:
                            phone_number = parts[0].strip()
                            time_zone = parts[1].strip()
                            
                            if phone_number:
                                # Create phone number with time zone for each selected sheet
                                for sheet in sheets:
                                    LeadPhoneNumbers.objects.get_or_create(
                                        lead=lead, 
                                        value=phone_number, 
                                        sheet=sheet,
                                        defaults={'time_zone': time_zone}
                                    )
                        else:
                            # Handle case where time zone is missing
                            phone_number = parts[0].strip()
                            if phone_number:
                                for sheet in sheets:
                                    LeadPhoneNumbers.objects.get_or_create(
                                        lead=lead, 
                                        value=phone_number, 
                                        sheet=sheet
                                    )
            
            # Process emails (no time zones)
            emails = form.cleaned_data.get('emails')
            if emails:
                for email in emails.split(','):
                    email = email.strip()
                    if email:
                        for sheet in sheets:
                            LeadEmails.objects.get_or_create(lead=lead, value=email, sheet=sheet)
            
            # Process contact names (no time zones)
            contact_names = form.cleaned_data.get('contact_names')
            if contact_names:
                for contact_name in contact_names.split(','):
                    contact_name = contact_name.strip()
                    if contact_name:
                        for sheet in sheets:
                            LeadContactNames.objects.get_or_create(lead=lead, value=contact_name, sheet=sheet)
            
            # Associate lead with all selected sheets
            for sheet in sheets:
                sheet.leads.add(lead)
            
            messages.success(request, "Lead added successfully!")
            
            if group_name == 'operations_manager':
                return redirect('operations_manager:index')
            elif group_name == 'operations_team_leader':
                return redirect('operations_team_leader:index')
            else:
                return HttpResponseForbidden("You do not have permission to perform this action.")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LeadForm()
    
    return render(request, "main/manage_lead_update.html", {'form': form})

@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def edit_lead(request, pk):
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name
    logger.debug(f"User group: {group_name}")

    lead = get_object_or_404(Lead, id=pk)

    if request.method == 'POST':
        form = LeadForm(request.POST, instance=lead)
        if form.is_valid():
            lead = form.save(commit=False)

            sheets = form.cleaned_data.get('sheets')
            if not sheets:
                messages.error(request, "Please select at least one sheet.")
                return render(request, 'main/manage_lead_update.html', {
                    'form': form,
                })

            lead.save()

            # Process phone numbers with time zones
            phone_data = form.cleaned_data.get('phone_numbers')
            new_phone_data = {}
            
            if phone_data:
                for line in phone_data.split('\n'):
                    line = line.strip()
                    if line:
                        # Split by comma, but be careful of time zones with commas
                        parts = line.split(',', 1)  # Split into max 2 parts
                        if len(parts) == 2:
                            phone_number = parts[0].strip()
                            time_zone = parts[1].strip()
                            new_phone_data[phone_number] = time_zone
                        else:
                            # If no time zone provided, use None
                            phone_number = parts[0].strip()
                            new_phone_data[phone_number] = None

            # Update Phone Numbers with Time Zones
            current_phone_objects = LeadPhoneNumbers.objects.filter(lead=lead)
            
            # Delete phone numbers that were removed
            current_phone_numbers = current_phone_objects.values_list('value', flat=True)
            numbers_to_delete = set(current_phone_numbers) - set(new_phone_data.keys())
            LeadPhoneNumbers.objects.filter(lead=lead, value__in=numbers_to_delete).delete()
            
            # Update existing or create new phone numbers with time zones
            for phone_number, time_zone in new_phone_data.items():
                if phone_number:  # Skip empty phone numbers
                    for sheet in sheets:
                        obj, created = LeadPhoneNumbers.objects.get_or_create(
                            lead=lead,
                            value=phone_number,
                            sheet=sheet,
                            defaults={'time_zone': time_zone}
                        )
                        if not created:
                            # Update time zone if it was provided
                            if time_zone is not None:
                                obj.time_zone = time_zone
                                obj.save()

            # Update Emails (no time zones)
            current_emails = LeadEmails.objects.filter(lead=lead).values_list('value', flat=True)
            new_emails = {email.strip() for email in form.cleaned_data.get('emails').split(',') if email.strip()}

            # Delete removed emails
            LeadEmails.objects.filter(lead=lead).exclude(value__in=new_emails).delete()
            # Add new emails
            for email in new_emails.difference(current_emails):
                for sheet in sheets:
                    LeadEmails.objects.get_or_create(lead=lead, value=email, sheet=sheet)

            # Update Contact Names (no time zones)
            current_contact_names = LeadContactNames.objects.filter(lead=lead).values_list('value', flat=True)
            new_contact_names = {contact_name.strip() for contact_name in form.cleaned_data.get('contact_names').split(',') if contact_name.strip()}

            # Delete removed contact names
            LeadContactNames.objects.filter(lead=lead).exclude(value__in=new_contact_names).delete()
            # Add new contact names
            for contact_name in new_contact_names.difference(current_contact_names):
                for sheet in sheets:
                    LeadContactNames.objects.get_or_create(lead=lead, value=contact_name, sheet=sheet)

            # Update Sheets
            current_sheets = lead.sheets.all()
            # Remove the lead from sheets that were deselected
            for sheet in current_sheets.difference(sheets):
                sheet.leads.remove(lead)
            # Add the lead to newly selected sheets
            for sheet in sheets.difference(current_sheets):
                sheet.leads.add(lead)

            messages.success(request, "Lead updated successfully!")

            if group_name == 'operations_manager':
                return redirect('operations_manager:index')
            elif group_name == 'operations_team_leader':
                return redirect('operations_team_leader:index')
            else:
                return HttpResponseForbidden("You do not have permission to perform this action.")
        else:
            # Log form and formset errors
            logger.debug("Form or formsets are invalid")
            messages.error(request, "Please correct the errors below.")
            for error in form.errors:
                messages.error(request, f"Form Error: {error}")
                logger.debug(f"Form Error: {error}")

    else:
        # Initialize form with existing lead data
        form = LeadForm(instance=lead)
        
        # Set initial values for other fields
        form.fields['sheets'].initial = lead.sheets.all()
        
        # Format phone numbers with time zones for display
        phone_numbers_with_time_zones = []
        for pn in LeadPhoneNumbers.objects.filter(lead=lead):
            if pn.time_zone:
                phone_numbers_with_time_zones.append(f"{pn.value},{pn.time_zone}")
            else:
                phone_numbers_with_time_zones.append(pn.value)
        
        form.fields['phone_numbers'].initial = '\n'.join(phone_numbers_with_time_zones)
        form.fields['emails'].initial = ', '.join(LeadEmails.objects.filter(lead=lead).values_list('value', flat=True))
        form.fields['contact_names'].initial = ', '.join(LeadContactNames.objects.filter(lead=lead).values_list('value', flat=True))

    return render(request, "main/manage_lead_update.html", {
        'form': form,
    })

@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def delete_lead(request, pk):
    # Fetch the lead instance or return a 404 error if not found
    lead = get_object_or_404(Lead, pk=pk)
    
    # Get the user's group name
    group_name = None
    if request.user.groups.exists():
        group_name = request.user.groups.first().name

    if request.method == 'POST':
        lead.sheets.clear()
        lead.delete()
        
        # Redirect based on user group
        if group_name == 'operations_manager':
            return redirect('operations_manager:index')
        elif group_name == 'operations_team_leader':
            return redirect('operations_team_leader:index')
        else:
            return HttpResponseForbidden("You do not have permission to perform this action.")

    # Determine the template path based on group
    if group_name == 'operations_manager':
        template_path = 'operations_manager/delete_lead.html'
    elif group_name == 'operations_team_leader':
        template_path = 'operations_team_leader/delete_lead.html'
    else:
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Render the template with lead information
    return render(request, template_path, {'lead': lead})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def auto_fill(request):

    template_name = 'operations_team_leader/auto_fill.html'  # Default template

    if request.user.groups.filter(name='operations_manager').exists():
        template_name = 'operations_manager/auto_fill.html'

    if request.method == 'POST':
        form = AutoFillForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            selected_sheet = form.cleaned_data.get('latest_sheet')

            try:
                # Load data
                if file.name.endswith('.xlsx'):
                    data = pd.read_excel(file, engine='openpyxl')
                elif file.name.endswith('.xls'):
                    data = pd.read_excel(file)
                elif file.name.endswith('.csv'):
                    data = pd.read_csv(file)
                else:
                    messages.error(request, "Unsupported file format.")
                    return render(request, template_name, {'form': form})

                # Ensure required columns are present
                required_columns = ['Company Name', 'Time Zone', 'Phone Number', 'Email', 'DM Name', 'Color']
                for col in required_columns:
                    if col not in data.columns:
                        data[col] = ''

                num_leads_total = len(data['Company Name'].unique())
                data['Company Name'] = data['Company Name'].map(clean_company_name)
                data = data[data['Company Name'].apply(filter_companies)]
                company_names = data['Company Name'].unique()

                leads = Lead.objects.filter(name__in=company_names)
                leads_dict = {lead.name: lead for lead in leads}
                num_leads_autofilled = len(leads)

                latest_leads_dict = {}

                # Only proceed with the latest sheet if it is selected
                if selected_sheet:
                    latest_leads = selected_sheet.leads.all()
                    latest_leads_dict = {lead.name: lead for lead in latest_leads}

                    # Add new rows to the data from the latest sheet
                    for lead in latest_leads:
                        phone_number, time_zone, email, contact_name = get_lead_related_data(lead)
                        data = data[data['Company Name'] != lead.name]

                        new_row = pd.DataFrame([{
                            'Company Name': lead.name,
                            'Time Zone': time_zone,
                            'Phone Number': phone_number,
                            'Email': email,
                            'DM Name': contact_name,
                            'Color': 'green'
                        }])

                        # Ensure the new_row has the same columns as data
                        new_row = new_row.reindex(columns=data.columns)

                        # Append new_row to the existing DataFrame
                        data = pd.concat([data, new_row], ignore_index=True)

                def fill_data(row):
                    company_name = row.get('Company Name')
                    if company_name not in latest_leads_dict:
                        lead = leads_dict.get(company_name)
                        if lead:
                            phone_number, time_zone, email, contact_name = get_lead_related_data(lead)
                            row['Time Zone'] = time_zone
                            row['Phone Number'] = phone_number
                            row['Email'] = email
                            row['DM Name'] = contact_name
                            row['Color'] = 'blue'
                        else:
                            row['Time Zone'] = ''
                            row['Phone Number'] = ''
                            row['Email'] = ''
                            row['DM Name'] = ''
                            row['Color'] = 'red'
                    return row

                data = data.apply(fill_data, axis=1)
                data.reset_index(drop=True, inplace=True)

                # Save file to server
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    data.to_excel(writer, index=False, sheet_name='Leads')
                    workbook = writer.book
                    worksheet = writer.sheets['Leads']

                    column_widths = {
                        'Company Name': 20,
                        'Time Zone': 15,
                        'Phone Number': 20,
                        'Email': 25,
                        'DM Name': 20,
                        'Color': 10
                    }

                    for col, width in column_widths.items():
                        col_idx = data.columns.get_loc(col) + 1
                        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

                    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

                    for idx, row in data.iterrows():
                        color = row['Color']
                        fill = green_fill if color == 'green' else blue_fill if color == 'blue' else red_fill
                        for col_num in range(1, len(row) + 1):
                            worksheet.cell(row=idx + 2, column=col_num).fill = fill

                output.seek(0)
                
                # Provide file download to the user
                filename = f"result_{file.name}"
                response = HttpResponse(
                    output,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                logger.info(f"[{request.user.username}] Made an AutoFill with [{num_leads_total}] total leads, [{num_leads_autofilled}] autofilled from the database.")
                Log.objects.create(message=f"[{request.user.username}] Made an AutoFill with [{num_leads_total}] total leads, [{num_leads_autofilled}] autofilled from the database.")

                return response

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                logger.error(f"Error processing file: {str(e)}", exc_info=True)
                return render(request, template_name, {'form': form})

    else:
        form = AutoFillForm()
        messages.error(request, "Form is not valid. Please check the form fields.")
        print(form.errors)  # Print form errors to terminal

    return render(request, template_name, {'form': form})


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def upload_sheet(request):
    template_name = "operations_team_leader/upload_sheet.html" if is_in_group(request.user, "operations_team_leader") else "operations_manager/upload_sheet.html"

    form = UploadSheetsForm()

    if request.method == 'POST':
        form = UploadSheetsForm(request.POST, request.FILES)
        if 'file' in request.FILES:
            file = request.FILES['file']
            
            # Save the file using the default storage (S3 or local)
            from django.core.files.storage import default_storage
            
            # Use default_storage to save the file. This works for both local and S3.
            # We use file.name as the path. S3 storage will handle it.
            # Local storage will save it to MEDIA_ROOT.
            file_name_in_storage = default_storage.save(file.name, file)
            
            # Call the Celery task with the storage file path/name
            process_sheet_task.delay(file_name_in_storage, file.name, request.user.id)

            messages.success(request, "Your sheet has been uploaded and is being processed in the background. You will be notified upon completion.")
            
            return render(request, template_name, {"form": form})
        else:
            messages.error(request, "Please upload a file.")
            return render(request, template_name, {"form": form})
    
    return render(request, template_name, {"form": form})


@user_passes_test(lambda user: is_in_group(user, "leads") or is_in_group(user, "operations_team_leader"))
def leads_average_view(request):
    user = request.user
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get the date range
    start_date_obj, end_date_obj = None, None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            start_date_obj = None

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            end_date_obj = None

    # Case for "leads" group
    if is_in_group(user, "leads"):
        leads_averages = LeadsAverage.objects.filter(user=user)
        if start_date_obj:
            leads_averages = leads_averages.filter(created_at__gte=start_date_obj)
        if end_date_obj:
            leads_averages = leads_averages.filter(created_at__lte=end_date_obj)

        # Calculate total average for the user
        total_leads_average = leads_averages.aggregate(Sum('count'))['count__sum'] or 0
        template_path = 'leads/leads_average.html'

        return render(request, template_path, {
            'leads_average': total_leads_average,
            'leads_averages': leads_averages,
            'start_date': start_date or '',
            'end_date': end_date or '',
        })

    # Case for "operations_team_leader" group
    elif is_in_group(user, "operations_team_leader"):
        # Get the team members of the leader (the user)
        team_members = UserLeader.objects.filter(leader=user).values_list('user', flat=True)
        team_leads_averages = LeadsAverage.objects.filter(user__in=team_members)
        leader_averages = LeadsAverage.objects.filter(user=user)

        if start_date_obj:
            team_leads_averages = team_leads_averages.filter(created_at__gte=start_date_obj)
            leader_averages = leader_averages.filter(created_at__gte=start_date_obj)

        if end_date_obj:
            team_leads_averages = team_leads_averages.filter(created_at__lte=end_date_obj)
            leader_averages = leader_averages.filter(created_at__lte=end_date_obj)

        # Calculate team averages
        team_averages = team_leads_averages.values('user').annotate(total_average=Sum('count'))
        user_averages = {}
        for avg in team_averages:
            user_id = avg['user']
            user = User.objects.get(id=user_id)
            user_averages[user.username] = avg['total_average']

        leader_total_average = leader_averages.aggregate(Sum('count'))['count__sum'] or 0
        total_team_average = leader_total_average + sum(user_averages.values())

        # Prepare data for rendering
        context = {
            'user_averages': {
                'team_members': user_averages,
                'leader': leader_total_average
            },
            'total_team_average': total_team_average,
            'leads_averages': team_leads_averages.union(leader_averages),
            'start_date': start_date or '',
            'end_date': end_date or '',
        }

        template_path = 'operations_team_leader/leads_average.html'
        return render(request, template_path, context)

    # Not authorized
    else:
        raise PermissionDenied("You are not authorized to view this page.")


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def manage_filter_words(request):

    template_name = "operations_team_leader/manage_filter_words.html" if is_in_group(request.user, "operations_team_leader") else "operations_manager/manage_filter_words.html"
    redirect_name = "operations_team_leader:manage-filter-words" if is_in_group(request.user, "operations_team_leader") else "operations_manager:manage-filter-words"

    query = request.GET.get('q', '')  # Get search query from the request
    filter_words_list = FilterWords.objects.all()

    if query:
        filter_words_list = filter_words_list.filter(word__icontains=query)  # Filter words by the search query

    filter_words_list = filter_words_list.order_by("-id")
    paginator = Paginator(filter_words_list, 30)
    page_number = request.GET.get('page')
    filter_words = paginator.get_page(page_number)

    if request.method == 'POST':
        form = FilterWordsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(redirect_name)
    else:
        form = FilterWordsForm()

    return render(request, template_name, {
        'filter_words': filter_words,
        'form': form,
        'query': query,  # Pass the query to the template to keep the search term
    })


@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def delete_filter_word(request, word_id):
    redirect_name = "operations_team_leader:manage-filter-words" if is_in_group(request.user, "operations_team_leader") else "operations_manager:manage-filter-words"
    filter_word = get_object_or_404(FilterWords, id=word_id)
    filter_word.delete()  # Delete the word
    return redirect(redirect_name)  # Redirect to the same page


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def log_inactivity(request):
    try:
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({"status": "failure", "message": "User not authenticated"}, status=401)
        
        user = request.user
        
        # Parse JSON data
        try:
            data = json.loads(request.body)
            message = data.get('message', 'User inactive')
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}")
            return JsonResponse({"status": "failure", "message": "Invalid JSON data"}, status=400)
        except Exception as e:
            logger.error(f"Error parsing request body: {e}")
            return JsonResponse({"status": "failure", "message": "Error parsing request data"}, status=400)
        
        # Log the inactivity
        try:
            # Check if user is in sales or sales_team_leader for SalesLog
            if user.groups.filter(name__in=['sales', 'sales_team_leader']).exists():
                # Create SalesLog entry
                SalesLog.objects.create(
                    message=message,
                    date=timezone.now(),
                    user=user
                )
            
            # Create general Log entry for all users
            Log.objects.create(
                message=f"{user.username} was inactive for 5 minutes"
            )
            
            logger.info(f"Inactivity logged for user: {user.username}")
            
            return JsonResponse({
                "status": "success", 
                "message": "Inactivity logged successfully",
                "user": user.username,
                "timestamp": timezone.now().isoformat()
            })
            
        except Exception as e:
            logger.error(f"Database error while logging inactivity: {e}")
            return JsonResponse({"status": "failure", "message": "Database error"}, status=500)
            
    except Exception as e:
        logger.error(f"Unexpected error in log_inactivity: {e}")
        return JsonResponse({"status": "failure", "message": "Server error"}, status=500)



@user_passes_test(lambda user: is_in_group(user, "sales_team_leader") or is_in_group(user, "sales_manager"))
def sales_log_view(request):
    # Get the current user
    user = request.user

    # Check if the user is in the 'sales_manager' group
    if user.groups.filter(name='sales_manager').exists():
        # If the user is a sales manager, filter logs for users in 'sales_team_leader' or 'sales' groups
        sales_logs = SalesLog.objects.filter(
            Q(user__groups__name='sales_team_leader') | Q(user__groups__name='sales')
        ).order_by("-id")
        logs = sales_logs  # Combine the two querysets
    else:
        # Otherwise, filter logs based on the user's team
        team_members = UserLeader.objects.filter(leader=user).values_list('user', flat=True)
        logs = SalesLog.objects.filter(user__in=team_members).order_by("-id")  # Filter logs for team members



    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        logs = logs.filter(
            Q(message__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # Date filter functionality
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if start_date and end_date:
        logs = logs.filter(date__range=[start_date, end_date])

    # Pagination
    paginator = Paginator(logs, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'logs': page_obj,
        'search_query': search_query,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'sales_team_leader/sales_logs.html', context)


from django.utils import timezone


def send_cb_date_notifications():
    now = timezone.now()
    task_name = 'send_cb_date_notifications'

    try:
        task_log = TaskLog.objects.get(task_name=task_name)
        if task_log.last_run and task_log.last_run.date() == now.date():
            # Task has already run today, no need to run it again
            return
    except TaskLog.DoesNotExist:
        task_log = TaskLog(task_name=task_name)

    # Use timezone-aware dates for the current day's range
    today_start = now.astimezone().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = now.astimezone().replace(hour=23, minute=59, second=59, microsecond=999999)

    flags = TerminationCode.objects.filter(name__in=['CB', 'PR']).all()
    leads_with_cb_today = LeadTerminationCode.objects.filter(CB_date__range=(today_start, today_end), flag__in=flags)

    if not leads_with_cb_today:
        logger.info("No leads found with CB_date today.")
        return  # Exit the function if no leads are found

    for lead_termination in leads_with_cb_today:
        user = lead_termination.user
        
        if user is None:
            logger.warning("Lead termination has no associated user. Skipping.")
            continue
        
        # Attempt to send notification to the user
        try:
            Notification.objects.create(
                sender=user,
                receiver=user,
                message=f"You have a call back due today for lead '{lead_termination.lead.name}' in show '{lead_termination.sales_show.name}', Please Check your Prospect sheet",
                notification_type=5
            )
        except Exception as e:
            logger.error(f"Failed to create notification for user {user.username}: {e}")
        
        # Notify the user's leader
        try:
            user_leader = UserLeader.objects.get(user=user).leader
            Notification.objects.create(
                sender=user,
                receiver=user_leader,
                message=f"{user.username} has a call back due today for lead '{lead_termination.lead.name}' in show '{lead_termination.sales_show.name}'",
                notification_type=5
            )
        except UserLeader.DoesNotExist:
            logger.info(f"No leader found for user: {user.username}. Skipping notification to leader.")
        except Exception as e:
            logger.error(f"Failed to create notification for leader {user_leader.username}: {e}")

    # Update last_run after processing all notifications
    task_log.last_run = now
    task_log.save()



@login_required
def mark_as_read(request, notification_id):
    from django.core.cache import cache

    role = request.user.groups.first().name

    notification = get_object_or_404(Notification, id=notification_id, receiver=request.user)
    notification.read = True
    notification.save()
    
    # Invalidate cache for unread notifications count
    cache_key = f'unread_notifications_{request.user.id}'
    cache.delete(cache_key)
    
    return redirect(request.META.get('HTTP_REFERER', reverse(f'{role}:notifications')))



## Function to import X shows and an earlier version of it was used to import the inventory 
@user_passes_test(lambda user: is_in_group(user, "operations_team_leader") or is_in_group(user, "operations_manager"))
def import_folder(request):
    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('files')
        uploaded_sheets = []
        skipped_files = []
        errors = []

        def process_file(file):
            try:
                # Read file into DataFrame
                data = None
                file_extension = str(file.name).split('.')[-1]
                if file_extension == 'xlsx':
                    data = pd.read_excel(file, engine='openpyxl', header=None)
                elif file_extension == 'xls':
                    data = pd.read_excel(file, header=None)
                elif file_extension == 'csv':
                    data = pd.read_csv(file, header=None)
                else:
                    skipped_files.append(file.name)
                    return None

                if data.empty:
                    skipped_files.append(file.name)
                    return None

                def ensure_str(value):
                    return str(value) if pd.notna(value) else ''

                expected_headers = ['Company Name', 'Phone Number', 'Time Zone', 'Email', 'DM Name']
                first_row_as_str = data.iloc[0].apply(ensure_str)
                has_header = all(header.lower() in [col.lower() for col in first_row_as_str if col] for header in expected_headers)

                if has_header:
                    data.columns = data.iloc[0].apply(ensure_str).tolist()
                    data = data[1:]  # Skip the header row

                data['Company Name'] = data['Company Name'].map(clean_company_name)
                data = data[data['Company Name'].apply(filter_companies)]

                # Ensure atomic transaction
                with transaction.atomic():
                    random_suffix = generate_random_string(5)  # Generate a random suffix
                    unique_sheet_name = f"{file.name}_{random_suffix}"  # Create a unique name
                    sheet = Sheet.objects.create(
                        name=unique_sheet_name,
                        user=request.user,
                        created_at=datetime.now()
                    )

                    for _, row in data.iterrows():
                        if not has_valid_contact(row):
                            continue

                        company_name = ensure_str(row.get('Company Name', ''))
                        lead, created = Lead.objects.get_or_create(name=company_name)
                        if created:
                            lead.save()
                            
                        time_zone = ensure_str(row.get('Time Zone', ''))

                        phone_numbers = filter(None, [ensure_str(row.get('Phone Number', '')).strip()])
                        for phone_number in phone_numbers:
                            if phone_number and is_valid_phone_number(phone_number):
                                LeadPhoneNumbers.objects.get_or_create(lead=lead, sheet=sheet, value=phone_number, time_zone=time_zone)

                        email = ensure_str(row.get('Email', ''))
                        if email:
                            LeadEmails.objects.get_or_create(lead=lead, sheet=sheet, value=email)

                        contact_name = ensure_str(row.get('DM Name', ''))
                        if contact_name:
                            LeadContactNames.objects.get_or_create(lead=lead, sheet=sheet, value=contact_name)

                        # Handle color if the 'Color' column exists
                        if 'Color' in data.columns:
                            color = get_string_value(row, 'Color')
                            if color and color.lower() in ['white', 'green', 'blue', 'red']:
                                LeadsColors.objects.create(lead=lead, sheet=sheet, color=color.lower())

                        if sheet not in lead.sheets.all():
                            lead.sheets.add(sheet)

                        sheet.is_approved = True
                        sheet.is_x = True   # Only for this function because it uploades x shows 
                        sheet.save()
                return sheet
            except Exception as e:
                errors.append((file.name, str(e)))
                return None

        # Process files sequentially
        for file in uploaded_files:
            result = process_file(file)
            if result:
                uploaded_sheets.append(result)

        total_leads = sum(sheet.leads.count() for sheet in uploaded_sheets)

        return render(request, 'main/upload_x_folder.html', {
            'success': f"Uploaded {total_leads} total leads from {len(uploaded_sheets)} sheets.",
            'skipped_files': skipped_files,
            'errors': errors,
        })

    return render(request, 'main/upload_x_folder.html')

######################################################################################################################################################################################

## Function to import the sales inventory
{
# def import_lead_termination_history(request):
#     form = ImportSheetsForm()

#     if request.method == 'POST':
#         form = ImportSheetsForm(request.POST)
#         if form.is_valid():
#             folder_path = form.cleaned_data['folder_path']
#             if os.path.isdir(folder_path):
#                 files = os.listdir(folder_path)
#                 skipped_files = []
#                 errors = []

#                 def process_file(file):
#                     data = None
#                     file_extension = str(file).split('.')[-1]
#                     file_path = os.path.join(folder_path, file)

#                     try:
#                         if file_extension == 'xlsx':
#                             data = pd.read_excel(file_path, engine='openpyxl', header=None)
#                         elif file_extension == 'xls':
#                             data = pd.read_excel(file_path, header=None)
#                         elif file_extension == 'csv':
#                             data = pd.read_csv(file_path, header=None)
#                         else:
#                             return None  # Skip unsupported file types
#                     except Exception as e:
#                         logger.error(f"Error reading file {file}: {e}")
#                         return None  # Skip files that cannot be read

#                     # Ensure we have the required columns, except 'CB Date' which is optional
#                     required_columns = ['Termination Code', 'Special Notes']
#                     data.columns = data.iloc[0].tolist()  # Assuming first row contains column names
#                     data = data[1:]  # Skip header row

#                     # If the file doesn't have required columns, skip it
#                     if not all(col in data.columns for col in required_columns):
#                         logger.warning(f"File {file} is missing required columns: {', '.join(required_columns)}")
#                         skipped_files.append(file)
#                         return None

#                     # Create OldShow object for the file
#                     old_show, _ = OldShow.objects.get_or_create(name=file)

#                     for _, row in data.iterrows():
#                         try:
#                             # Fetch the lead from the database
#                             lead_name = row.get('Company Name', '').strip()
#                             lead = Lead.objects.filter(name__iexact=lead_name).first()
#                             if not lead:
#                                 logger.warning(f"Lead not found: {lead_name}")
#                                 continue

#                             # Fetch the TerminationCode object, skip if not found
#                             termination_code_str = row.get('Termination Code', '').strip()
#                             termination_code = TerminationCode.objects.filter(name=termination_code_str).first()
#                             if not termination_code:
#                                 logger.warning(f"Termination code not found: {termination_code_str} for lead {lead_name}. Skipping this lead.")
#                                 continue

#                             # CB Date is optional, so handle it separately
#                             cb_date = pd.to_datetime(row.get('CB Date', None), errors='coerce') if 'CB Date' in row else None

#                             # Special Notes
#                             notes = row.get('Special Notes', '').strip()

#                             # Create LeadTerminationHistory entry
#                             LeadTerminationHistory.objects.create(
#                                 lead=lead,
#                                 termination_code=termination_code,
#                                 cb_date=cb_date,
#                                 notes=notes,
#                                 old_show=old_show  # Save the OldShow reference
#                             )

#                         except Exception as e:
#                             logger.error(f"Error processing row: {row} - {e}")
#                             errors.append((file, str(e)))

#                     return old_show

#                 # Process files in parallel
#                 with ThreadPoolExecutor(max_workers=4) as executor:
#                     results = list(executor.map(process_file, files))

#                 # Log results and errors
#                 logger.info(f"Skipped {len(skipped_files)} files: {skipped_files}")
#                 logger.info(f"Encountered {len(errors)} errors: {errors}")

#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "success": "Sheets imported successfully."
#                 })
#             else:
#                 logger.error("The specified directory does not exist.")
#                 return render(request, "leads/import_sheet.html", {
#                     "form": form,
#                     "error": "The specified directory does not exist."
#                 })

#     return render(request, "leads/import_sheet.html", {
#         "form": form
#     })
}
